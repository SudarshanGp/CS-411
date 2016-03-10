__author__ = 'Nathan'
#!/usr/bin/env python
import timeit #test how long it takes
import os #removing old database
import math
import random
import sqlite3 #for the database
import time #for dealing with time objects
import xlrd #for reading .xlsx files
import xlsxwriter
from dateutil.parser import parse
import csv #for readin .csv files
from datetime import datetime, timedelta, time, date #for dealing with datetime objects
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
import sys
import os

db_name = 'Demo.db'
CurrentDir= "C:\\Users\\Nathan\\Desktop\\"

from HVAC_DB_Schema_Constants import *
from DB_Interface_08_28_15 import *


def AcademicCollege(fileloc):
    
    
    AcademicCollege=[]
    ReturnList=[]
    wb=xlrd.open_workbook(fileloc)
    current_sheet=wb.sheet_by_index(0)
    for i in range(5, 16): #starts at row 5 
        AcademicCollege=AcademicCollege+[current_sheet.cell(i,1).value]
    AcademicCollege.insert(0,str(sys.argv[1][0:4]))
    ReturnList.append(AcademicCollege)
    return ReturnList
    
def Gender(fileloc):
    
    
    Gender=[]
    ReturnList=[]
    wb=xlrd.open_workbook(fileloc)
    current_sheet=wb.sheet_by_index(0)
    for i in range(30, 34): #starts at row 5
        if i!=32: #we don't want the blank
            Gender=Gender+[current_sheet.cell(i,5).value]
    Gender.insert(0,str(sys.argv[1][0:4]))
    ReturnList.append(Gender)
    return ReturnList

def AcademicParser():
    fileloc=CurrentDir+"FreshmanData\\"+str(sys.argv[1])
    AcademicCollegeList=AcademicCollege(fileloc)
    print AcademicCollegeList
    buildOutputDatabase(db_name)
    conn = sqlite3.connect(db_name)
    writeAcademicCollege(conn, AcademicCollegeList)
    conn.commit()
    conn.close()

    return AcademicCollegeList

def GenderParser():
    fileloc=CurrentDir+"FreshmanData\\"+str(sys.argv[1])
    GenderList=Gender(fileloc)
    print GenderList
    buildOutputDatabase(db_name)
    conn = sqlite3.connect(db_name)
    writeGender(conn, GenderList)
    conn.commit()
    conn.close()

    return GenderList

def length(i):
    """returns the length of i"""
    return len(str(i))

def main():
    AcademicCollege=AcademicParser()
    Gender=GenderParser()
   
   
if __name__ == '__main__':
    #main()
    #runs the main, and prints the time taken to run
    print timeit.timeit(main,number=1)
