__author__ = 'Nathan'
#!/usr/bin/env python
import timeit #test how long it takes
import os #removing old database
import math
import random
import sqlite3 #for the database
import time #for dealing with time objects
import xlrd #for reading .xlsx files
import xlsxwriter
from dateutil.parser import parse
import csv #for readin .csv files
from datetime import datetime, timedelta, time, date #for dealing with datetime objects
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook
from bisect import bisect
import datetime, time
import sys
import os
import collections
import pickle
db_name = 'HVAC_ProccessedDB_vLatestall_11-17-15.db'
CurrentDir= "F:\\CBITEC DATA\\"

from HVAC_DB_Schema_Constants import *
from DB_Interface_08_28_15 import *


def validatorEcuData(datalist):
    newlist=[]
    errorlist=[None, None]
    for i in range(0, len(datalist)):
        if i==0: #ecupower
            try:
                float(datalist[i])
                if float(datalist[i])>MAXECUPOWER or float(datalist[i]) <0.0: #0<avgECUPower<15000 KWATTS
                    errorlist[i]=datalist[i]
                    datalist[i]=None
                else:
                    errorlist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    datalist[i]=None
                    errorlist[i]=None
                else:
                    print 'Was in i==0', datalist[i]
        if i==1: #ecumode
            if datalist[i]==None:
                errorlist[i]==None
            else:
                try:
                    str(datalist[i])
                    errorlist[i]=None
                except ValueError:
                    errorlist[i]=datalist[i]
                    datalist[i]=None



        #have to insert like this because other items still need to be added
    newlist.insert(0,datalist[0])
    newlist.insert(1,datalist[1])
    newlist.insert(2,errorlist[0])
    newlist.insert(3,errorlist[1])
    return newlist

def validatorShelterData(datalist):
    newlist=[]
    #print len(datalist)
    errorlist=[None,None,None]
    for i in range(0, len(datalist)):
        if i==0: #interior temp
            try:
                float(datalist[i])
                if float(datalist[i])>MAXAVGINTERIROAIRTEMP or float(datalist[i]) <-50: #0<avgECUPower<15000 KWATTS
                    errorlist[i]=datalist[i]
                    datalist[i]=None
                else:
                    errorlist[i]=None
            except ValueError and TypeError:
                print type(datalist[i])
                print repr(datalist[i])
                if datalist[i]=='' or datalist[i]=='---':
                    datalist[i]=None
                    errorlist[i]=None
                else:
                    print 'Was in i==0', datalist[i]
        if i==1: #avg  plug load
            try:
                float(datalist[i])
                if float(datalist[i])>MAXAVGPLUGLOAD or float(datalist[i]) <0.0: #0<avgECUPower<15000 KWATTS
                    errorlist[i]=datalist[i]
                    datalist[i]=None
                else:
                    errorlist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    datalist[i]=None
                    errorlist[i]=None
                else:
                    print 'Was in i==1', datalist[i]
        if i==2: #avgoccupancy
            try:
                float(datalist[i])
                if float(datalist[i])>1000000 or float(datalist[i]) <0: #0<avgECUPower<15000 KWATTS
                    errorlist[i]=datalist[i]
                    datalist[i]=None
                else:
                    errorlist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    datalist[i]=None
                    errorlist[i]=None
                else:
                    print 'Was in i==2', datalist[i]


        #have to insert like this because other items still need to be added
    newlist.insert(0,datalist[0])
    newlist.insert(1,datalist[1])
    newlist.insert(2,datalist[2])
    newlist.insert(3,errorlist[0])
    newlist.insert(4,errorlist[1])
    newlist.insert(5,errorlist[2])
    return newlist


def validatorWeatherStationData(datalist):
    newlist=[]
    errorlist=[None, None,None, None]
    for i in range(0, len(datalist)):
        #dont need to check time, if its empty you shouldnt even be reading that cell... also it cant be wrong... cant use flags as its stored in datetime

        if i==0:
            try:
                float(datalist[i])
                if float(datalist[i])>MAXAMBIENTAIRTEMP or float(datalist[i]) <-50.0: #20<ambientairtemp<115 F
                    errorlist[i]=datalist[i]
                    datalist[i]=None  #set it to negative numbers? so each item has its flag code
                else:
                    errorlist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    datalist[i]=None
                    errorlist[i]=None

                else:
                    print 'Was in i==1', datalist[i]
        elif i==1:
            try:
                float(datalist[i])
                if float(datalist[i]) <0.0 or float(datalist[i]) >MAXGLOBALHORIZONTALSOLAR: #0<globalHorizontalSolar<2000 WATTS per meter^2
                    errorlist[i]=datalist[i]
                    datalist[i]=None
                else:
                    errorlist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    datalist[i]=None
                    errorlist[i]=None
                else:
                #checks for ',' in the case '1,234.0' removes comma and then checks expression
                    try:
                        datalist[i]=datalist[i].replace(",","")
                        if float(datalist[i])<0.0 or float(datalist[i])>2000.0:
                            errorlist[i]=datalist[i]
                            datalist[i]=None
                    except ValueError:
                            print 'Was in i==2', datalist[i]
        elif i==2:
            try:
                float(datalist[i])
                if float(datalist[i])<1.5 or float(datalist[i])>MAXRELATIVEHUMIDTY: #0<relativeHumidity<100  percent
                    errorlist[i]=datalist[i]
                    datalist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    errorlist[i]=None
                    datalist[i]=None
                else:
                    print 'Was in i==3', datalist[i]
        elif i==3:
            try:
                float(datalist[i])
                if float(datalist[i])<0.0 or float(datalist[i])>MAXWINDSPEED: #0<windspeed<120 miles/hour
                    errorlist[i]=datalist[i]
                    datalist[i]=None
            except ValueError:
                if datalist[i]=='' or datalist[i]=='---':
                    errorlist[i]=None
                    datalist[i]=None
                else:
                    print 'Was in i==4', datalist[i]

    #have to insert like this because other items still need to be added
    newlist.insert(0,datalist[0])
    newlist.insert(1,datalist[1])
    newlist.insert(2,datalist[2])
    newlist.insert(3,datalist[3])
    newlist.insert(4,errorlist[0])
    newlist.insert(5,errorlist[1])
    newlist.insert(6,errorlist[2])
    newlist.insert(7,errorlist[3])
    return newlist



def ShelterEcuDataReader_DEMO_ONE_B_HUT(final_path_2,day,timetemp):
    count=0
    shelterlist=[]
    eculist=[]
    #shelterinstance2=['NPPR-0006_1','NPPR-0007_1','EB-0080_1','NPPR-0008_1','NPPR-0009_1','NPPR-0010_1','EB-0080_2','NPPR-0011_1']
    shelterinstance2=['NPPR-0006_1','EB-0080_1','EB-0080_2','NPPR-0008_1','NPPR-0009_1','NPPR-0010_1','EB-0080_3','NPPR-0011_1']
    weatherlist=[]
    datetimerownum=0
    ambienttemp=1
    windspeed=2
    relativehum=3
    solarrad=4
    Ecucolnum=[]
    Tempcolnum=[]
    for i in range(13, 77, 8):
        Ecucolnum.append(i)
        Tempcolnum.append(i-3)
    with open(final_path_2, 'rb') as f:
        reader = csv.reader(f)
        dateobj=datetime.datetime.strptime(day,"%Y%m%d").date()
        for row in reader:
            count+=1
            if count!=1:
                if row[windspeed]=="":
                    temp=0
                else:
                    temp=1
                Eculist=[]
                interiorTemplist=[]
                time1=row[datetimerownum][timetemp:20]
                timeobj=datetime.datetime.strptime(str(time1),"%H:%M:%S").time()
                Datetime_obj=datetime.datetime.combine(dateobj, timeobj)
                for a in range(0,len(Ecucolnum)):
                    Eculist.append(row[(Ecucolnum[a]+temp)])
                    interiorTemplist.append(row[(Tempcolnum[a]+temp)])
                    #print interiorTemplist

                AmbientTemp=row[ambienttemp]
                Windspeed=row[windspeed]
                Humidity=row[(relativehum+temp)]
                SolarHorz=row[(solarrad+temp)]
                for i in range(0,8):

                    if interiorTemplist[i]!="" and interiorTemplist[i]!="0" and interiorTemplist[i]!='0.00':
                        newlist=validatorShelterData([interiorTemplist[i],'',''])
                        shelterlist.append((["CBITEC_CBITEC-B-Hut_"+shelterinstance2[i],"CBITEC_CBITEC-B-Hut_NPPR-0001_1","SLB-STO-D_B-Hut",Datetime_obj,"1"]+newlist))
                    if Eculist[i]!="" and Eculist[i]!="0" and Eculist[i]!="0.00":
                        newlist2=validatorEcuData([Eculist[i],None])
                        eculist.append((["CBITEC_CBITEC-B-Hut_"+shelterinstance2[i], "CBITEC_CBITEC-B-Hut_NPPR-0001_1", "SLB-STO-D_B-Hut", Datetime_obj, "1"]+newlist2))
                    if AmbientTemp!="" and AmbientTemp!="0" and AmbientTemp!="0.00":
                        flag=1
                if  flag==1:
                    newlist3=validatorWeatherStationData([AmbientTemp,SolarHorz,Humidity,Windspeed])
                    weatherlist.append((["SLB-STO-D",Datetime_obj,"1"]+newlist3))
                flag=0

    return eculist,shelterlist, weatherlist


def CBITEECWeatherCSV(fileloc):
    weatherlist=[]
    wb=xlrd.open_workbook(fileloc)
    current_sheet=wb.sheet_by_index(0)
    for i in range(1, current_sheet.nrows):
        Value_Date_WeatherTime=current_sheet.cell(i,0).value
        Value_Time_WeatherTime=current_sheet.cell(i,1).value
        WeatherTemp=current_sheet.cell(i,2).value
        Humdity=current_sheet.cell(i,3).value
        windspeed=current_sheet.cell(i,4).value
        globalhoz=current_sheet.cell(i,7).value
        year1, month1, day1, hour, minute, second= xlrd.xldate_as_tuple(Value_Date_WeatherTime, wb.datemode)
        year, month, day, hour1, minute1, second1 = xlrd.xldate_as_tuple(Value_Time_WeatherTime, wb.datemode)
        if length(hour1) == 1:
            hour1= '0'+str(hour1)
        elif length(hour1) == 0:
            hour1= '00'+str(hour1)

        if length(minute1) == 1:
            minute1= '0'+str(minute1)
        elif length(minute1) == 0:
            minute1= '00'+str(minute1)

        if length(second1) == 1:
            second1= '0'+str(second1)
        elif length(second1) == 0:
            second1= '00'+str(second1)

        if length(month1) == 1:
            month1= '0'+str(month1)

        if length(day1) == 1:
            day1= '0'+str(day1)

        time1=str(hour1)+':'+str(minute1)+':'+str(second1)
        date=str(year1)+'-'+str(month1)+'-'+str(day1)
        strdatetime= date+' '+time1
        WeatherTemp_datetime=datetime.datetime(*time.strptime(strdatetime, "%Y-%m-%d %H:%M:%S")[:6])
        Weatherdata=validatorWeatherStationData([WeatherTemp, globalhoz, Humdity, windspeed])
        weatherlist.append(["DENT",WeatherTemp_datetime,"5"]+Weatherdata)
    return weatherlist

def CBITECShelterEcu(filepath):
    print 'inCBITECECU'
    eculist=[]

    shelterlist=[]
    datarow_start=2
    col_Plug_load_Breaker= 2
    col_Date_Breaker= 0
    col_Time_Breaker= 1
    col_tempavg_Breaker= 3
    Breakerdict={}
    col_Date_IECU=0
    col_Time_IECU=1
    col_ECU_Power_IECU=2
    IECUdict={}

    #shelterinstance2=['NPPR-0006_1','NPPR-0007_1','EB-0080_1','NPPR-0008_1','NPPR-0009_1','NPPR-0010_1','EB-0080_2','NPPR-0011_1']
    shelterinstance2=['NPPR-0006_1','EB-0080_1','EB-0080_2','NPPR-0008_1','NPPR-0009_1','NPPR-0010_1','EB-0080_3','NPPR-0011_1']



    #B-Hut_A NPPR-0006
    #B-Hut_B NPPR-0007
    #B-Hut_C EB-0080
    #B-Hut_D NPPR-0008
    #B-Hut_E NPPR-0009
    #B-Hut_F NPPR-0010
    #B-Hut_G EB-0080
    #B-Hut_H NPPR-0011
    wb = xlrd.open_workbook(filepath)
    for i in range(0, 16, 2):
        print i

        for j in range(0,2):
            print 'here:', j
            current_sheet= wb.sheet_by_index(i+j)
            if j==0:
                Value_Ecu_IECUlist=current_sheet.col_values(col_ECU_Power_IECU)

            if j==1:
                Value_Plug_Load_Breakerlist=current_sheet.col_values(col_Plug_load_Breaker)
                Value_Temp_Breakerlist=current_sheet.col_values(col_tempavg_Breaker)

            for k in range(datarow_start, current_sheet.nrows):
                if j==0:
                    #Do  the IECU STUFF HERE
                    Value_Date_IECU=current_sheet.cell(k,col_Date_IECU).value
                    Value_Time_IECU=current_sheet.cell(k,col_Time_IECU).value
                    year1, month1, day1, hour, minute, second= xlrd.xldate_as_tuple(Value_Date_IECU, wb.datemode)

                    year, month, day, hour1, minute1, second1 = xlrd.xldate_as_tuple(Value_Time_IECU, wb.datemode)
                    if length(hour1) == 1:
                        hour1= '0'+str(hour1)
                    elif length(hour1) == 0:
                        hour1= '00'+str(hour1)

                    if length(minute1) == 1:
                        minute1= '0'+str(minute1)
                    elif length(minute1) == 0:
                        minute1= '00'+str(minute1)

                    if length(second1) == 1:
                        second1= '0'+str(second1)
                    elif length(second1) == 0:
                        second1= '00'+str(second1)

                    if length(month1) == 1:
                        month1= '0'+str(month1)

                    if length(day1) == 1:
                        day1= '0'+str(day1)

                    time1=str(hour1)+':'+str(minute1)+':'+str(second1)
                    date=str(year1)+'-'+str(month1)+'-'+str(day1)
                    strdatetime= date+' '+time1
                    IECU_datetime=datetime.datetime(*time.strptime(strdatetime, "%Y-%m-%d %H:%M:%S")[:6])
                    temp=validatorEcuData([Value_Ecu_IECUlist[k],""])
                    #IECUdict[IECU_datetime]=[Value_Ecu_IECUlist[k]] #sets datetime-> ecu_power
                    if i==0:
                        rank=1
                    elif i==2:
                        rank=2
                    elif i==4:
                        rank=3
                    elif i==6:
                        rank=4
                    elif i==8:
                        rank=5
                    elif i==10:
                        rank=6
                    elif i==12:
                        rank=7
                    elif i==14:
                        rank=8
                    #print "B-Hut_"+str(rank)
                    #eculist.append(["B-Hut_"+str(rank),"I_ECU","DENT_B-Hut",IECU_datetime,"5"]+temp)
                    eculist.append(["CBITEC_CBITEC-B-Hut_"+shelterinstance2[rank-1],"CBITEC_CBITEC-B-Hut_NPPR-0001_1","Dent_B-Hut",IECU_datetime,"5"]+temp)

                elif j==1:
                    #Do the BREAKER STUFF HERE

                    Value_Date_Breaker=current_sheet.cell(k,col_Date_Breaker).value
                    Value_Time_Breaker=current_sheet.cell(k,col_Time_Breaker).value

                    year12, month12, day12, hour11, minute11, second11= xlrd.xldate_as_tuple(Value_Date_Breaker, wb.datemode)
                    year11, month11, day11, hour12, minute12, second12 = xlrd.xldate_as_tuple(Value_Time_Breaker, wb.datemode)

                    if length(hour12) == 1:
                        hour12= '0'+str(hour12)
                    elif length(hour12) == 0:
                        hour12= '00'+str(hour12)

                    if length(minute12) == 1:
                        minute12= '0'+str(minute12)
                    elif length(minute12) == 0:
                        minute12= '00'+str(minute12)

                    if length(second12) == 1:
                        second12= '0'+str(second12)
                    elif length(second12) == 0:
                        second12= '00'+str(second12)

                    if length(month12) == 1:
                        month12= '0'+str(month12)

                    if length(day12) == 1:
                        day12= '0'+str(day12)

                    time12=str(hour12)+':'+str(minute12)+':'+str(second12)
                    date12=str(year12)+'-'+str(month12)+'-'+str(day12)
                    strdatetime12= date12+' '+time12
                    temp2=validatorShelterData([Value_Temp_Breakerlist[k],Value_Plug_Load_Breakerlist[k],""])
                    #print temp2
                    Breaker_datetime=datetime.datetime(*time.strptime(strdatetime12, "%Y-%m-%d %H:%M:%S")[:6])
                    if i==0:
                        rank2=1
                    elif i==2:
                        rank2=2
                    elif i==4:
                        rank2=3
                    elif i==6:
                        rank2=4
                    elif i==8:
                        rank2=5
                    elif i==10:
                        rank2=6
                    elif i==12:
                        rank2=7
                    elif i==14:
                        rank2=8
                    #print "B-Hut_"+str(rank)
                    #shelterlist.append(["B-Hut_"+str(rank2),"I_ECU","Dent_B-Hut",Breaker_datetime,"5"]+temp2)
                    shelterlist.append(["CBITEC_CBITEC-B-Hut_"+shelterinstance2[rank2-1],"CBITEC_CBITEC-B-Hut_NPPR-0001_1","Dent_B-Hut",Breaker_datetime,"5"]+temp2)

                else:
                    print'ERROR!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'


    return eculist, shelterlist
def CBITEECECUWeatherDMMS(filelocationecuweather):
    #shelterinstance2=['NPPR-0006_1','NPPR-0007_1','EB-0080_1','NPPR-0008_1','NPPR-0009_1','NPPR-0010_1','EB-0080_2','NPPR-0011_1']
    shelterinstance2=['NPPR-0006_1','EB-0080_1','EB-0080_2','NPPR-0008_1','NPPR-0009_1','NPPR-0010_1','EB-0080_3','NPPR-0011_1']

    eculist=[]
    weatherlist=[]
    EcuA=2
    PlugA=4
    TempA=5
    EcuB=7
    PlugB=8
    TempB=10
    EcuC=12
    PlugC=13
    TempC=15
    EcuD=17
    PlugD=18
    TempD=20
    EcuE=22
    PlugE=23
    TempE=25
    EcuF=27
    PlugF=28
    TempF=30
    EcuG=32
    PlugG=33
    TempG=35
    EcuH=37
    PlugH=38
    TempH=40
    flag=0
    shelterlist=[]
    onlyfiles = [ f for f in listdir(filelocationecuweather) if isfile(join(filelocationecuweather,f))]
    print onlyfiles
    #onlyfiles=onlyfiles[6:7]
    for i in onlyfiles:
        count=0
        dulicates=[]

        print i
        with open(filelocationecuweather+i, 'rb') as f:
            reader = csv.reader(f)
            for row in reader:
                count+=1
                if count!=1:
                    Eculist=[]
                    Pluglist=[]
                    interiorTemplist=[]
                    Datetime = row[0]
                    ECUA = row[EcuA]
                    PLUGA = row[PlugA]
                    TEMPA = row[TempA]
                    ECUB = row[EcuB]
                    PLUGB = row[PlugB]
                    TEMPB = row[TempB]
                    ECUC = row[EcuC]
                    PLUGC = row[PlugC]
                    TEMPC = row[TempC]
                    ECUD = row[EcuD]
                    PLUGD = row[PlugD]
                    TEMPD = row[TempD]
                    ECUE = row[EcuE]
                    PLUGE = row[PlugE]
                    TEMPE = row[TempE]
                    ECUF = row[EcuF]
                    PLUGF = row[PlugF]
                    TEMPF = row[TempF]
                    ECUG = row[EcuG]
                    PLUGG = row[PlugG]
                    TEMPG = row[TempG]
                    ECUH = row[EcuH]
                    PLUGH = row[PlugH]
                    TEMPH = row[TempH]
                    Eculist.append([ECUA,ECUB,ECUC,ECUD,ECUE,ECUF,ECUG,ECUH])
                    Pluglist.append([PLUGA,PLUGB,PLUGC,PLUGD,PLUGE,PLUGF,PLUGG,PLUGH])
                    interiorTemplist.append([TEMPA,TEMPB,TEMPC,TEMPD,TEMPE,TEMPF,TEMPG,TEMPH])
                    AmbientTemp = row[41]
                    Humidity=row[42]
                    WindSpeed=row[43]
                    Barometricpressure=row[44]
                    d = datetime.datetime.strptime(Datetime,"%m/%d/%Y %H:%M")
                    Datetime=d.strftime("%Y-%m-%d %H:%M:%S")
                    Datetime_obj=datetime.datetime(*time.strptime(Datetime, "%Y-%m-%d %H:%M:%S")[:6])
                    for i in range(0,8):
                        if interiorTemplist[0][i]!="" and interiorTemplist[0][i]!="0" and interiorTemplist[0][i]!="0.00":
                            if Pluglist[0][i]==""or Pluglist[0][i]=="0" or Pluglist[0][i]=="0.00":
                                Pluglist[0][i]=0.0
                            newlist=validatorShelterData([interiorTemplist[0][i],Pluglist[0][i],''])
                            shelterlist.append((["CBITEC_CBITEC-B-Hut_"+shelterinstance2[i],"CBITEC_CBITEC-B-Hut_NPPR-0001_1","DMMS_B-Hut",Datetime_obj,"5"]+newlist))

                        if Eculist[0][i]!="" and Eculist[0][i]!="0" and Eculist[0][i]!="0.00":
                            newlist2=validatorEcuData([Eculist[0][i],None])
                            eculist.append((["CBITEC_CBITEC-B-Hut_"+shelterinstance2[i], "CBITEC_CBITEC-B-Hut_NPPR-0001_1", "DMMS_B-Hut", Datetime_obj, "5"]+newlist2))


                        if AmbientTemp!="" and AmbientTemp!="0" and AmbientTemp!="0.00":
                            flag=1
                if  flag==1:
                    newlist3=validatorWeatherStationData([AmbientTemp,Barometricpressure,Humidity,WindSpeed])
                    weatherlist.append((["DMMS",Datetime_obj,"5"]+newlist3))
                flag=0

    return eculist, shelterlist, weatherlist

def BHut_SLB():
    """Imports  SLB-STO-D B-Hut data at CBITEC"""
    fileloc=CurrentDir+"\\SLB-STO-D\\"
    files = [ f for f in listdir(fileloc) if isfile(join(fileloc,f))]
    ecufinal=[]
    shelterfinal=[]
    weatherfinal=[]
    for i in range(0,len(files)):
        day=str(files[i][47:55])
        print day
        final_path_2=fileloc+files[i]
        Eculist, Shelterlist, Weatherlist = ShelterEcuDataReader_DEMO_ONE_B_HUT(final_path_2,day,12)
        ecufinal.extend(Eculist)
        shelterfinal.extend(Shelterlist)
        weatherfinal.extend(Weatherlist)
    return ecufinal, shelterfinal, weatherfinal

def CBITEC_DMMS():
    """Imports DMMS B-Hut data at CBITEC"""
    filelocationecuweather=CurrentDir+"CBITEC\\DMMS\\"
    Eculist, Shelterlist, Weatherlist=CBITEECECUWeatherDMMS(filelocationecuweather)
    return Eculist, Shelterlist, Weatherlist

def CBITEC_DENT():
    """Imports Dent B-Hut data at CBITEC"""
    filelocationecu=CurrentDir+"CBITEC\\CBITEC_EcuData_reduced.xlsx"
    filelocationweather2=CurrentDir+"CBITEC\\CBITEC_WeatherData.xlsx"
    Weatherlist=CBITEECWeatherCSV(filelocationweather2)
    Eculist, Shelterlist=CBITECShelterEcu(filelocationecu)
    return Eculist, Shelterlist, Weatherlist

def OptimizeWeather(Wl1, wl2, wl3):
    #w1l 4-13-15 -> 4-23-15 --> dont use
    #wl2 4-1-15 -> 4-30-15
    #    12-1-14 -> 12-31-14
    #    2-1-15 ->  2-6-15
    #    1-1-15 ->  1-6-15
    #    1-26-15 -> 1-31-15
    #    7-15-14 -> 7-31-14 --> dont use
    #    3-9-15 -> 3-31-15
    #    11-1-14 ->  11-5-14
    #    11-18-14 -> 11-25-14
    #    10-1-14  -> 10-23-14
    #    10-30-14 -> 10-31-14
    #    9-9-14  -> 9-30-14
    #wl3 2-28-14 -> 2-28-14
    #    3-1-14 -> 8-18-14
    finalweather=[]
    for i in wl2:
        #print type(i[1])
        if(datetime.date(2014,7,15)<=i[1].date() and i[1].date()<=datetime.date(2014,7,31)):
            pass
        elif(datetime.datetime(2015,4,23,15,35,0)<=i[1] and i[1]<=datetime.datetime(2015,4,30,23,55,0)):
            pass
        elif(datetime.datetime(2015,3,16,14,55,0)<=i[1] and i[1]<=datetime.datetime(2015,3,17,0,15,0)):
            pass
        elif(datetime.datetime(2015,3,18,11,30,0)<=i[1] and i[1]<=datetime.datetime(2015,3,24,20,0,0)):
            pass
        elif(datetime.datetime(2014,11,19,3,55,0)<=i[1] and i[1]<=datetime.datetime(2014,11,25,6,20,0)):
            pass
        elif(datetime.datetime(2015,2,1,0,0,0)<=i[1] and i[1]<=datetime.datetime(2015,2,5,9,0,0)):
            pass
        elif(datetime.datetime(2014,10,1,0,0,0)<=i[1] and i[1]<=datetime.datetime(2014,10,7,6,40,0)):
            pass
        else:
            finalweather.append(i)
    for j in wl3:
        finalweather.append(j)
    return finalweather

def  OptimizeEcu(Eculist1,Eculist2,Eculist3):
    #w1l 4-13-15 -> 4-23-15
    #wl2 4-1-15 -> 4-30-15  --> dont use this!!!! bad values and not all huts
    #    12-1-14 -> 12-31-14
    #    2-1-15 ->  2-6-15
    #    1-1-15 ->  1-6-15
    #    1-26-15 -> 1-31-15
    #    7-15-14 -> 7-31-14                 ------------
    #    3-9-15 -> 3-31-15
    #    11-1-14 ->  11-5-14
    #    11-18-14 -> 11-25-14
    #    10-1-14  -> 10-23-14
    #    10-30-14 -> 10-31-14
    #    9-9-14  -> 9-30-14
    #wl3 2-28-14 -> 2-28-14 dent
    #    3-1-14 -> 8-18-14                   --------------
    finalecu=[]
    fix=[]
    fix2=[]
    ShelterId=['NPPR-0006','NPPR-0007','EB-0080','NPPR-0008','NPPR-0009','NPPR-0010','EB-0080','NPPR-0011']
    for k in Eculist3:
        #compare these later
        if(datetime.date(2014,7,15)<=k[3].date() and k[3].date()<=datetime.date(2014,7,31)):
            fix.append(k)
            pass
        else:
            finalecu.append(k)
    for i in Eculist2:
        #bad values to avoid
        if(datetime.date(2014,4,1)<=i[3].date() and i[3].date()<=datetime.date(2014,4,30)):
            pass
        #compare these later
        elif(datetime.date(2014,7,15)<=i[3].date() and i[3].date()<=datetime.date(2014,7,31)):
            fix2.append(i)
        else:
            finalecu.append(i)
    for j in Eculist1:
        finalecu.append(j)


    #now take care of (2014,7,15 -- 2014,7,31)
    print len(fix2)
    print len(fix)
    starttime=datetime.datetime(2014,7,15,6,5,0)#07/15/2014 06:05:00
    endtime=datetime.datetime(2014,7,30,17,0,0) #07/30/2014 17:00:00
    curtime=starttime
    check=check2=check3=0
    counter=0
    while(1):
        counter+=1
        if counter>1:
            curtime=curtime+datetime.timedelta(minutes=5)
        for j in ShelterId:
            for k in fix:
                if k[0]==j and k[3]==curtime:
                    for a in fix2:
                        if a[0]==k[0] and a[3]==k[3]:
                            if(a[7]==None and k[7]==None):
                                newlist=[a[0],a[1],a[2],a[3],a[4],((float(a[5])+float(k[5]))/2),a[6],(a[7]),a[8]]
                            elif(a[7]==None and k[7]!=None):
                                newlist=[a[0],a[1],a[2],a[3],a[4],float(a[5]),a[6],a[7],a[8]]
                            elif(a[7]!=None and k[7]==None):
                                newlist=[a[0],a[1],a[2],a[3],a[4],float(k[5]),a[6],k[7],a[8]]
                            else:
                                newlist=[a[0],a[1],a[2],a[3],a[4],a[5],a[6],((a[7]+k[7])/2),a[8]]
                            #if comes in here two data points
                            #average the data
                            finalecu.append(newlist)
                            #append on the final list
                            check=1
                    if check != 1:
                        finalecu.append(k)
                    if check != 1:
                        check2=1
            if check!=1 and check2!=1:
                #double check fix2 may have not been in fix
                for y in fix2:
                    if y[0]==j and y[3]==curtime:
                        finalecu.append(y)
                        #only in fix2
                        #append on the list
                        check3=1
            if check!=1 and check2!=1 and check3!=1:
                #that hut is not there at that time period
                pass
        check=0
        check2=0
        check3=0
        if(curtime==endtime):
            print 'finished Ecu check'
            break


    return finalecu

    # book = xlsxwriter.Workbook("Eculist.xlsx")
    # sheet1 = book.add_worksheet()
    # count=0
    # for i in range(0,len(Eculist1)):
    #     count+=1
    #     for j in range(0,len(Eculist1[0])):
    #         if j==3:
    #             Eculist1[i][j]=Eculist1[i][j].strftime('%m/%d/%Y %H:%M:%S')
    #         sheet1.write(count,j,Eculist1[i][j])
    # i=0
    # j=0
    # for i in range(0,len(Eculist2)):
    #     count+=1
    #     for j in range(0,len(Eculist2[0])):
    #         if j==3:
    #             Eculist2[i][j]=Eculist2[i][j].strftime('%m/%d/%Y %H:%M:%S')
    #         sheet1.write(count,j,Eculist2[i][j])
    # i=0
    # j=0
    # for i in range(0,len(Eculist3)):
    #     count+=1
    #     for j in range(0,len(Eculist3[0])):
    #        if j==3:
    #            Eculist3[i][j]=Eculist3[i][j].strftime('%m/%d/%Y %H:%M:%S')
    #        sheet1.write(count,j,Eculist3[i][j])
    # book.close()
def OptimizeShelter(Shelterlist1,Shelterlist2,Shelterlist3):
    #w1l 4-13-15 -> 4-23-15
    #wl2 4-1-15 -> 4-30-15  --> dont use this!!!! bad values and not all huts
    #    12-1-14 -> 12-31-14
    #    2-1-15 ->  2-6-15
    #    1-1-15 ->  1-6-15
    #    1-26-15 -> 1-31-15
    #    7-15-14 -> 7-31-14                 ------------>>  DONT USE THIS
    #    3-9-15 -> 3-31-15
    #    11-1-14 ->  11-5-14
    #    11-18-14 -> 11-25-14
    #    10-1-14  -> 10-23-14
    #    10-30-14 -> 10-31-14
    #    9-9-14  -> 9-30-14
    #wl3 2-28-14 -> 2-28-14 dent
    #    3-1-14 -> 8-18-14
    finalshelter=[]
    for k in Shelterlist3:
            finalshelter.append(k)
    for i in Shelterlist2:
        #bad values to avoid
        if((datetime.date(2014,4,1)<=i[3].date() and i[3].date()<=datetime.date(2014,4,30)) or (datetime.date(2014,7,15)<=i[3].date() and i[3].date()<=datetime.date(2014,7,31)) ):
            pass
        else:
            finalshelter.append(i)
    for j in Shelterlist1:
        finalshelter.append(j)
    #################################OUTPUTS VALUES INTO AN EXCEL FILE#####################3
    # book = xlsxwriter.Workbook("Shelterlist.xlsx")
    # sheet1 = book.add_worksheet()
    # count=0
    # for i in range(0,len(Shelterlist1)):
    #     count+=1
    #     for j in range(0,len(Shelterlist1[0])):
    #         if j==3:
    #             Shelterlist1[i][j]=Shelterlist1[i][j].strftime('%m/%d/%Y %H:%M:%S')
    #         sheet1.write(count,j,Shelterlist1[i][j])
    # i=0
    # j=0
    # for i in range(0,len(Shelterlist2)):
    #     count+=1
    #     for j in range(0,len(Shelterlist2[0])):
    #         if j==3:
    #            Shelterlist2[i][j]=Shelterlist2[i][j].strftime('%m/%d/%Y %H:%M:%S')
    #         sheet1.write(count,j,Shelterlist2[i][j])
    # i=0
    # j=0
    # for i in range(0,len(Shelterlist3)):
    #     count+=1
    #     for j in range(0,len(Shelterlist3[0])):
    #         if j==3:
    #             Shelterlist3[i][j]=Shelterlist3[i][j].strftime('%m/%d/%Y %H:%M:%S')
    #         sheet1.write(count,j,Shelterlist3[i][j])
    # book.close()
    return finalshelter

def Databaseimport(finalshelter,finalecu, finalweather):
    """Imports the three lists into the database"""
    LocationID ='CBITEC'
    Shelter_info=[['NPPR-0006','ADD DESCRIPTION'],['NPPR-0008','ADD DESCRIPTION'],['NPPR-0009','ADD DESCRIPTION'],['NPPR-0010','ADD DESCRIPTION'],['NPPR-0011','ADD DESCRIPTION'],['EB-0080','ADD DESCRIPTION']]
    Ecu_info=[['NPPR-0001','ADD DESCRIPTION']]
    ShelterInstance_info1=[['CBITEC_CBITEC-B-Hut_NPPR-0006_1','NPPR-0006',LocationID,'NPPR-0006 add description'],['CBITEC_CBITEC-B-Hut_EB-0080_1','EB-0080',LocationID,'EB-0080 add description'],['CBITEC_CBITEC-B-Hut_EB-0080_2','EB-0080',LocationID,'B-Hut Baseline Hut'],['CBITEC_CBITEC-B-Hut_NPPR-0008_1','NPPR-0008',LocationID,'NPPR-0008 add description'],['CBITEC_CBITEC-B-Hut_NPPR-0009_1','NPPR-0009',LocationID,'NPPR-0009 add description'],['CBITEC_CBITEC-B-Hut_NPPR-0010_1','NPPR-0010',LocationID,'NPPR-0010 add description'],['CBITEC_CBITEC-B-Hut_EB-0080_3','EB-0080',LocationID,'EB-0080 add description'],['CBITEC_CBITEC-B-Hut_NPPR-0011_1','NPPR-0011',LocationID,'NPPR-0011 add description']]
    EcuInstance_info1=[['CBITEC_CBITEC-B-Hut_NPPR-0001_1','NPPR-0001',LocationID,'NPPR-0001 add description']]


    #--prob need more than one weather station and location Id---#
    #--same location: CBITEC
    weatherStationID = 'DMMS'
    weatherStationDescription ='CBITEC_B-HUT add weather description'
    dataSetID1 = 'Dent_B-Hut'
    datagroupId1='B-Hut_DataGroup'
    sensors1='ADD Sensor description here'
    dataSetDescription1 = 'CBITEC Dent add description'
    DocumenatationDes1= 'Add Documentation Description here'
    weather_station_info = [[weatherStationID,LocationID, weatherStationDescription]]
    data_set_info1 = [[dataSetID1,datagroupId1,LocationID,dataSetDescription1,sensors1,DocumenatationDes1]]

    weatherStationID2 ='DENT'
    weatherStationDescription2 ='CBITEC_B-HUT add weather description'
    dataSetID2 = 'DMMS_B-Hut'
    datagroupId2='B-Hut_DataGroup'
    sensors2='ADD Sensor description here'
    dataSetDescription2 = 'CBITEC DMMS add description'
    DocumenatationDes2= 'Add Documentation Description here'
    weather_station_info2 = [[weatherStationID2,LocationID, weatherStationDescription2]]
    data_set_info2 = [[dataSetID2, datagroupId2,LocationID,dataSetDescription2,sensors2, DocumenatationDes2]]

    weatherStationID3= 'SLB-STO-D'
    weatherStationDescription3 ='CBITEC_B-HUT add weather description'
    dataSetID3=  'SLB-STO-D_B-Hut'
    datagroupId3='B-Hut_DataGroup'
    sensors3='ADD Sensor description here'
    dataSetDescription3 = 'Add datasetDescription here'
    DocumenatationDes3= 'Add Documentation Description here'
    weather_station_info3 = [[weatherStationID3,LocationID, weatherStationDescription3]]
    data_set_info3 = [[dataSetID3, datagroupId3, LocationID,dataSetDescription3,sensors3, DocumenatationDes3]]


    print 'Loading Database'
    initializeHVACDB(db_name)
    conn = sqlite3.connect(db_name)
    writeDataGroupInfo(conn,[["B-Hut_DataGroup","Collection of B-Hut datasets"]])
    writeLocationInfo(conn, [[LocationID,"Add Location Description"]])
    writeECUTypeInfo(conn,None,Ecu_info)
    writeShelterTypeInfo(conn,None, Shelter_info)
    writeWeatherStationInfo(conn, weather_station_info) #DENT
    writeWeatherStationInfo(conn, weather_station_info2) #DMMS
    writeWeatherStationInfo(conn, weather_station_info3) #SLB-STO-D
    writeDataSetInfo(conn,data_set_info1) #DENT
    writeDataSetInfo(conn,data_set_info2) #DMMS
    writeDataSetInfo(conn,data_set_info3) #SLB-STO-D
    writeECUInstanceInfo(conn, EcuInstance_info1)
    writeShelterInstanceInfo(conn, ShelterInstance_info1)
    writeShelterData(conn, finalshelter)
    writeWeatherData(conn, finalweather)
    writeECUData(conn, finalecu)
    conn.commit()
    conn.close()


def length(i):
    """returns the length of i"""
    return len(str(i))

def main():
    #-----Goes through the B-Hut datasets and seperates into three lists
    #no need to uncomment, just use Pickle Load!!!
    #Eculist1, Shelterlist1, Weatherlist1 = BHut_SLB()
    #Eculist2, Shelterlist2, Weatherlist2 = CBITEC_DMMS()
    #Eculist3, Shelterlist3, Weatherlist3 = CBITEC_DENT()

    #-----Only, uncomment when you make a chance to the above three functions!!!
    #pickle.dump(Weatherlist1, open("PickleFiles\\w1.p","wb"))
    #pickle.dump(Weatherlist2, open("PickleFiles\\w2.p","wb"))
    #pickle.dump(Weatherlist3, open("PickleFiles\\w3.p","wb"))
    #pickle.dump(Eculist1, open("PickleFiles\\e1.p","wb"))
    #pickle.dump(Eculist2, open("PickleFiles\\e2.p","wb"))
    #pickle.dump(Eculist3, open("PickleFiles\\e3.p","wb"))
    #pickle.dump(Shelterlist1, open("PickleFiles\\s1.p","wb"))
    #pickle.dump(Shelterlist2, open("PickleFiles\\s2.p","wb"))
    #pickle.dump(Shelterlist3, open("PickleFiles\\s3.p","wb"))

    #-----loads the Weatherlists from the three different datasets
    #Weatherlist1=pickle.load(open("PickleFiles\\w1.p","rb"))
    #Weatherlist2=pickle.load(open("PickleFiles\\w2.p","rb"))
    #Weatherlist3=pickle.load(open("PickleFiles\\w3.p","rb"))

    #-----Takes the three lists and makes one combine accurate Weatherlist over the time span of all the three lists
    #-----Saves the list in the FinalWeather.p file
    #-----only uncomment the bottom two lines if you want to make an edit to the Optimize Weather Function
    #finalweather=OptimizeWeather(Weatherlist1,Weatherlist2, Weatherlist3)
    #pickle.dump(finalweather, open("PickleFiles\\FinalWeather.p","wb"))

    #-----loads the Eculist from the three different datasets
    #Eculist1=pickle.load(open("PickleFiles\\e1.p","rb"))
    #Eculist2=pickle.load(open("PickleFiles\\e2.p","rb"))
    #Eculist3=pickle.load(open("PickleFiles\\e3.p","rb"))

    #-----Takes the three lists and makes one combine accurate Weatherlist over the time span of all the three lists
    #-----Saves the list in the FinalWeather.p file
    #-----only uncomment the bottom two lines if you want to make an edit to the Optimize Weather Function
    #finalecu=OptimizeEcu(Eculist1,Eculist2,Eculist3) # ----> make sure filter algo works
    #pickle.dump(finalecu, open("PickleFiles\\FinalEcu.p","wb"))

    #-----loads the Shelterlist from the three different datasets
    #Shelterlist1=pickle.load(open("PickleFiles\\s1.p","rb"))
    #Shelterlist2=pickle.load(open("PickleFiles\\s2.p","rb"))
    #Shelterlist3=pickle.load(open("PickleFiles\\s3.p","rb"))

    #-----Takes the three lists and makes one combine accurate Weatherlist over the time span of all the three lists
    #-----Saves the list in the FinalWeather.p file
    #-----only uncomment the bottom two lines if you want to make an edit to the Optimize Weather Function
    #finalshelter=OptimizeShelter(Shelterlist1,Shelterlist2,Shelterlist3)
    #pickle.dump(finalshelter, open("PickleFiles\\FinalShelter.p","wb"))

    #-----takes all the final lists and imports them into the database
    finalshelter=pickle.load(open("PickleFiles\\FinalShelter.p","rb"))
    finalecu=pickle.load(open("PickleFiles\\FinalEcu.p","rb"))
    finalweather=pickle.load(open("PickleFiles\\FinalWeather.p","rb"))
    print 'Finished Loading'
    Databaseimport(finalshelter,finalecu, finalweather)


if __name__ == '__main__':
    #main()
    #runs the main, and prints the time taken to run
    print timeit.timeit(main,number=1)
